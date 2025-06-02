import psutil
import datetime
import os
from openpyxl import Workbook, load_workbook

# Excel file name
excel_file = "daily_data_usage.xlsx"

# Get current date
today = datetime.date.today().strftime("%Y-%m-%d")

# Get total bytes sent and received
net_io = psutil.net_io_counters()
total_bytes = net_io.bytes_sent + net_io.bytes_recv
total_mb = round(total_bytes / (1024 ** 2), 2)  # Convert to MB

# Create or update Excel file
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Total Data Used (MB)"])
else:
    wb = load_workbook(excel_file)
    ws = wb.active

# Avoid duplicate entries for the same day
dates = [row[0].value for row in ws.iter_rows(min_row=2)]
if today not in dates:
    ws.append([today, total_mb])

wb.save(excel_file)

print(f"{today}: {total_mb} MB logged.")
